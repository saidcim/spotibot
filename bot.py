"""
Spotify AI Playlist Bot
- Her 2 günde bir dinleme geçmişini analiz eder, AI ile playlist günceller.
- Flask API ile Vercel dashboard'a endpoint sağlar.
- State Railway environment variable + /data volume'da saklanır.
"""

import os
import json
import time
import logging
import datetime
import threading
import openpyxl
import schedule
import spotipy
import requests
from spotipy.oauth2 import SpotifyOAuth
from groq import Groq
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
log = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

# ── Config ─────────────────────────────────────────────────────────────────
SPOTIFY_CLIENT_ID     = os.environ["SPOTIFY_CLIENT_ID"]
SPOTIFY_CLIENT_SECRET = os.environ["SPOTIFY_CLIENT_SECRET"]
SPOTIFY_REDIRECT_URI  = os.environ.get("SPOTIFY_REDIRECT_URI", "http://127.0.0.1:5000/callback")
SPOTIFY_REFRESH_TOKEN = os.environ["SPOTIFY_REFRESH_TOKEN"]
GROQ_API_KEY          = os.environ["GROQ_API_KEY"]
API_SECRET            = os.environ.get("API_SECRET", "said")
PLAYLIST_NAME         = os.environ.get("PLAYLIST_NAME", "🤖 AI Daily Mix")
PLAYLIST_SIZE         = int(os.environ.get("PLAYLIST_SIZE", "40"))

RAILWAY_API_TOKEN      = os.environ.get("RAILWAY_API_TOKEN", "")
RAILWAY_PROJECT_ID     = os.environ.get("RAILWAY_PROJECT_ID", "")
RAILWAY_SERVICE_ID     = os.environ.get("RAILWAY_SERVICE_ID", "")
RAILWAY_ENVIRONMENT_ID = os.environ.get("RAILWAY_ENVIRONMENT_ID", "")

DATA_DIR = os.environ.get("DATA_DIR", "/data" if os.path.isdir("/data") else ".")
os.makedirs(DATA_DIR, exist_ok=True)
HISTORY_FILE = os.path.join(DATA_DIR, "playlist_history.xlsx")
STATE_FILE   = os.path.join(DATA_DIR, "bot_state.json")
STATE_VAR    = "BOT_STATE"

ARCHIVE_LIMIT = 50

# carry_over artık sabit değil — AI, geri bildirim trendine bakarak
# state["dynamic_config"]["carry_over"] içindeki değeri önerir (bkz. tune_dynamic_config()).
# playlist_size ise SABİT kalır (PLAYLIST_SIZE env değişkeninden).
DEFAULT_CARRY_OVER = 5
CARRY_OVER_MIN, CARRY_OVER_MAX = 0, 10

is_running = False


def _default_user_profile() -> dict:
    return {
        "version": 1,
        "last_updated": None,
        "learned_patterns": [],
        "genre_affinity": {},
        "artist_affinity": {},
        "mood_music_map": {},
        "next_advice": "",
        "adaptation_metrics": {
            "cycles_completed": 0,
            "avg_score_first_3": 0.0,
            "avg_score_last_3": 0.0,
            "avg_plays_trend": "0%",
            "carry_over_success_rate": 0.0,
            "confidence": 0.0,
            "improvement_delta": 0.0,
        },
    }


def _default_dynamic_config() -> dict:
    return {
        "carry_over": DEFAULT_CARRY_OVER,
        "last_tuned_cycle": 0,
        "tune_reason": "",
    }


def _default_state() -> dict:
    return {
        "playlist_id": os.environ.get("PLAYLIST_ID", ""),
        "last_update": None,
        "cycle": 0,
        "feedback_history": [],
        "mood_history": [],
        "playlist_archive": [],
        "user_profile": _default_user_profile(),
        "dynamic_config": _default_dynamic_config(),
        "ai_notes": "",
    }


# ── State ────────────────────────────────────────────────────────────────────

def _railway_headers():
    return {"Authorization": f"Bearer {RAILWAY_API_TOKEN}", "Content-Type": "application/json"}


def _can_use_railway():
    return all([RAILWAY_API_TOKEN, RAILWAY_PROJECT_ID, RAILWAY_SERVICE_ID, RAILWAY_ENVIRONMENT_ID])


def _load_state_from_file(path: str) -> dict | None:
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.warning(f"State dosyası okunamadı ({path}): {e}")
        return None


def _merge_state_defaults(state: dict) -> dict:
    default = _default_state()
    for key, val in default.items():
        if key not in state:
            state[key] = val
    if "user_profile" in state:
        profile_default = _default_user_profile()
        for k, v in profile_default.items():
            if k not in state["user_profile"]:
                state["user_profile"][k] = v
        metrics = profile_default["adaptation_metrics"]
        for k, v in metrics.items():
            if k not in state["user_profile"].get("adaptation_metrics", {}):
                state["user_profile"]["adaptation_metrics"][k] = v

    dyn_default = _default_dynamic_config()
    dyn = state.get("dynamic_config") or {}
    for k, v in dyn_default.items():
        if k not in dyn:
            dyn[k] = v
    # Güvenlik: state bozulmuş/elle değiştirilmiş olsa bile sınırların dışına çıkmasın
    dyn["carry_over"] = max(CARRY_OVER_MIN, min(CARRY_OVER_MAX, int(dyn["carry_over"])))
    dyn["carry_over"] = min(dyn["carry_over"], PLAYLIST_SIZE)  # carry_over playlist_size'ı geçemez
    state["dynamic_config"] = dyn
    return state


def load_state() -> dict:
    state = None

    if _can_use_railway():
        try:
            query = """
            query Variables($projectId: String!, $serviceId: String!, $environmentId: String!) {
              variables(projectId: $projectId, serviceId: $serviceId, environmentId: $environmentId)
            }"""
            resp = requests.post(
                "https://backboard.railway.app/graphql/v2",
                headers=_railway_headers(),
                json={"query": query, "variables": {
                    "projectId": RAILWAY_PROJECT_ID,
                    "serviceId": RAILWAY_SERVICE_ID,
                    "environmentId": RAILWAY_ENVIRONMENT_ID,
                }},
                timeout=10,
            )
            data = resp.json()
            variables = data.get("data", {}).get("variables", {})
            if STATE_VAR in variables:
                state = json.loads(variables[STATE_VAR])
                log.info(f"State Railway'den yüklendi (döngü #{state.get('cycle', 0)})")
        except Exception as e:
            log.warning(f"Railway state okunamadı: {e}")

    file_state = _load_state_from_file(STATE_FILE)
    if file_state is None:
        file_state = _load_state_from_file("bot_state.json")

    if state is None:
        state = file_state or _default_state()
    elif file_state and file_state.get("cycle", 0) > state.get("cycle", 0):
        log.info("Dosyadaki state daha güncel, dosya kullanılıyor.")
        state = file_state

    return _merge_state_defaults(state)


def save_state(state: dict):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)

    if _can_use_railway():
        try:
            mutation = """
            mutation UpsertVariables($input: VariableCollectionUpsertInput!) {
              variableCollectionUpsert(input: $input)
            }"""
            requests.post(
                "https://backboard.railway.app/graphql/v2",
                headers=_railway_headers(),
                json={"query": mutation, "variables": {"input": {
                    "projectId": RAILWAY_PROJECT_ID,
                    "serviceId": RAILWAY_SERVICE_ID,
                    "environmentId": RAILWAY_ENVIRONMENT_ID,
                    "variables": {STATE_VAR: json.dumps(state, ensure_ascii=False)},
                }}},
                timeout=10,
            )
            log.info("State Railway'e kaydedildi.")
        except Exception as e:
            log.warning(f"Railway state kaydedilemedi: {e}")


# ── Excel Geçmişi ──────────────────────────────────────────────────────────

TRACK_HEADERS = ["Döngü", "Tarih", "Saat", "Spotify ID", "Şarkı", "Sanatçı", "Albüm", "AI Skoru", "Çalma Sayısı"]
SUMMARY_HEADERS = ["Döngü", "Tarih-Saat", "Ruh Hali", "Enerji", "Skor", "Şarkı Sayısı",
                   "Ort. Çalma", "Taşınan Şarkı", "AI Analiz Özeti"]


def _get_or_create_sheets(wb):
    if "Şarkılar" in wb.sheetnames:
        ws_tracks = wb["Şarkılar"]
    else:
        ws_tracks = wb.active
        ws_tracks.title = "Şarkılar"
    if "Döngü Özeti" in wb.sheetnames:
        ws_summary = wb["Döngü Özeti"]
    else:
        ws_summary = wb.create_sheet("Döngü Özeti")
    if ws_tracks.max_row == 1 and ws_tracks.cell(1, 1).value is None:
        for col, h in enumerate(TRACK_HEADERS, 1):
            ws_tracks.cell(1, col).value = h
    if ws_summary.max_row == 1 and ws_summary.cell(1, 1).value is None:
        for col, h in enumerate(SUMMARY_HEADERS, 1):
            ws_summary.cell(1, col).value = h
    return ws_tracks, ws_summary


def save_to_excel(tracks: list, cycle: int, score=None, play_counts: dict | None = None):
    play_counts = play_counts or {}
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    wb = openpyxl.load_workbook(HISTORY_FILE) if os.path.exists(HISTORY_FILE) else openpyxl.Workbook()
    ws_tracks, _ = _get_or_create_sheets(wb)

    for track in tracks:
        plays = play_counts.get(track["id"], 0)
        ws_tracks.append([
            cycle, date_str, time_str, track["id"], track["name"],
            track["artist"], track["album"], score if score is not None else "—", plays,
        ])
    wb.save(HISTORY_FILE)
    log.info(f"Excel'e {len(tracks)} şarkı kaydedildi (döngü {cycle})")


def save_cycle_summary_to_excel(cycle: int, mood_data: dict, score: float, track_count: int,
                                 avg_plays: float, carry_count: int, analysis: str):
    now = datetime.datetime.now()
    datetime_str = now.strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.load_workbook(HISTORY_FILE) if os.path.exists(HISTORY_FILE) else openpyxl.Workbook()
    _, ws_summary = _get_or_create_sheets(wb)
    ws_summary.append([
        cycle, datetime_str,
        mood_data.get("mood", "—"),
        mood_data.get("energy_level", "—"),
        score, track_count, round(avg_plays, 2), carry_count,
        (analysis or "")[:200],
    ])
    wb.save(HISTORY_FILE)
    log.info(f"Excel döngü özeti kaydedildi (döngü {cycle})")


# ── Spotify ────────────────────────────────────────────────────────────────

def get_spotify() -> spotipy.Spotify:
    auth = SpotifyOAuth(
        client_id=SPOTIFY_CLIENT_ID,
        client_secret=SPOTIFY_CLIENT_SECRET,
        redirect_uri=SPOTIFY_REDIRECT_URI,
        scope=(
            "user-read-recently-played user-top-read user-library-read "
            "playlist-modify-public playlist-modify-private playlist-read-private"
        ),
    )
    token_info = auth.refresh_access_token(SPOTIFY_REFRESH_TOKEN)
    return spotipy.Spotify(auth=token_info["access_token"])


def get_listening_data(sp: spotipy.Spotify) -> dict:
    data = {}
    recent = sp.current_user_recently_played(limit=50)
    recent_tracks = []
    cutoff = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=3)
    for item in recent["items"]:
        played_at = datetime.datetime.strptime(
            item["played_at"], "%Y-%m-%dT%H:%M:%S.%fZ"
        ).replace(tzinfo=datetime.timezone.utc)
        if played_at >= cutoff:
            t = item["track"]
            recent_tracks.append({
                "id": t["id"], "name": t["name"],
                "artist": t["artists"][0]["name"], "album": t["album"]["name"],
                "played_at": item["played_at"],
            })
    data["recent_tracks"] = recent_tracks

    top_short = sp.current_user_top_tracks(limit=50, time_range="short_term")
    data["top_short"] = [{"id": t["id"], "name": t["name"], "artist": t["artists"][0]["name"],
        "popularity": t["popularity"]} for t in top_short["items"]]

    top_medium = sp.current_user_top_tracks(limit=50, time_range="medium_term")
    data["top_medium"] = [{"id": t["id"], "name": t["name"], "artist": t["artists"][0]["name"],
        "popularity": t["popularity"]} for t in top_medium["items"]]

    top_artists = sp.current_user_top_artists(limit=20, time_range="short_term")
    data["top_artists"] = [{"name": a["name"], "genres": a["genres"]} for a in top_artists["items"]]

    saved = sp.current_user_saved_tracks(limit=50)
    data["saved_tracks"] = [{"id": item["track"]["id"], "name": item["track"]["name"],
        "artist": item["track"]["artists"][0]["name"]} for item in saved["items"]]

    return data


def get_playlist_play_counts(playlist_track_ids, recent_tracks):
    counts = {tid: 0 for tid in playlist_track_ids}
    for t in recent_tracks:
        if t["id"] in counts:
            counts[t["id"]] += 1
    return counts


def build_listening_fingerprint(listening_data: dict) -> dict:
    genres = []
    for a in listening_data.get("top_artists", [])[:10]:
        genres.extend(a.get("genres", [])[:3])
    genre_counts: dict[str, int] = {}
    for g in genres:
        genre_counts[g] = genre_counts.get(g, 0) + 1
    top_genres = sorted(genre_counts, key=genre_counts.get, reverse=True)[:5]
    return {
        "top_artists": [a["name"] for a in listening_data.get("top_artists", [])[:8]],
        "top_genres": top_genres,
        "track_count_3d": len(listening_data.get("recent_tracks", [])),
    }


def find_or_create_playlist(sp: spotipy.Spotify, state: dict) -> str:
    user_id = sp.current_user()["id"]
    if state.get("playlist_id"):
        try:
            pl = sp.playlist(state["playlist_id"])
            log.info(f"Mevcut playlist: {pl['name']} ({pl['id']})")
            return state["playlist_id"]
        except Exception:
            log.warning("Playlist ID geçersiz, aranıyor...")

    matching = []
    offset = 0
    while True:
        results = sp.current_user_playlists(limit=50, offset=offset)
        for pl in results["items"]:
            if pl["name"] == PLAYLIST_NAME and pl["owner"]["id"] == user_id:
                matching.append(pl)
        if results["next"] is None:
            break
        offset += 50

    if matching:
        keeper = matching[0]
        for dup in matching[1:]:
            try:
                sp.current_user_unfollow_playlist(dup["id"])
                log.info(f"Kopya silindi: {dup['id']}")
            except Exception as e:
                log.warning(f"Kopya silinemedi: {e}")
        return keeper["id"]

    pl = sp.user_playlist_create(
        user=user_id, name=PLAYLIST_NAME, public=False,
        description="🤖 Her 2 günde güncellenen AI playlist",
    )
    log.info(f"Yeni playlist: {pl['id']}")
    return pl["id"]


def update_playlist(sp, playlist_id, track_ids):
    sp.playlist_replace_items(playlist_id, [])
    for i in range(0, len(track_ids), 100):
        chunk = [f"spotify:track:{tid}" for tid in track_ids[i:i + 100]]
        sp.playlist_add_items(playlist_id, chunk)
    log.info(f"Playlist güncellendi: {len(track_ids)} şarkı")


# ── Öğrenme ────────────────────────────────────────────────────────────────

def analyze_patterns(state: dict, listening_data: dict, play_counts: dict,
                     current_track_ids: list) -> dict:
    profile = state.setdefault("user_profile", _default_user_profile())
    archive = state.get("playlist_archive", [])
    feedback = state.get("feedback_history", [])
    fingerprint = build_listening_fingerprint(listening_data)

    genre_affinity: dict[str, float] = dict(profile.get("genre_affinity", {}))
    artist_affinity: dict[str, float] = dict(profile.get("artist_affinity", {}))
    mood_music_map: dict[str, list] = dict(profile.get("mood_music_map", {}))

    for genre in fingerprint["top_genres"]:
        genre_affinity[genre] = min(1.0, genre_affinity.get(genre, 0.0) + 0.1)
    for artist in fingerprint["top_artists"][:5]:
        artist_affinity[artist] = min(1.0, artist_affinity.get(artist, 0.0) + 0.1)

    if archive and feedback:
        last_archive = archive[-1]
        last_mood = last_archive.get("mood", "")
        if last_mood and fingerprint["top_genres"]:
            existing = mood_music_map.get(last_mood, [])
            for g in fingerprint["top_genres"][:3]:
                if g not in existing:
                    existing.append(g)
            mood_music_map[last_mood] = existing[:5]

    carry_limit = state.get("dynamic_config", {}).get("carry_over", DEFAULT_CARRY_OVER)
    carry_plays = sum(play_counts.get(tid, 0) for tid in current_track_ids[:carry_limit])
    total_plays = sum(play_counts.values()) or 1
    carry_rate = carry_plays / total_plays if current_track_ids else 0.0

    scores = [f["score"] for f in feedback if "score" in f]
    avg_first_3 = sum(scores[:3]) / len(scores[:3]) if scores else 0.0
    avg_last_3 = sum(scores[-3:]) / len(scores[-3:]) if scores else 0.0
    improvement = round(avg_last_3 - avg_first_3, 2) if len(scores) >= 3 else 0.0

    plays_trend = "0%"
    if len(archive) >= 2:
        early = sum(a.get("avg_plays", 0) for a in archive[:3]) / min(3, len(archive))
        recent = sum(a.get("avg_plays", 0) for a in archive[-3:]) / min(3, len(archive))
        if early > 0:
            pct = int((recent - early) / early * 100)
            plays_trend = f"{pct:+d}%"

    profile["genre_affinity"] = genre_affinity
    profile["artist_affinity"] = artist_affinity
    profile["mood_music_map"] = mood_music_map
    profile["adaptation_metrics"] = {
        "cycles_completed": state.get("cycle", 0),
        "avg_score_first_3": round(avg_first_3, 2),
        "avg_score_last_3": round(avg_last_3, 2),
        "avg_plays_trend": plays_trend,
        "carry_over_success_rate": round(carry_rate, 2),
        "confidence": profile.get("adaptation_metrics", {}).get("confidence", 0.0),
        "improvement_delta": improvement,
    }
    profile["last_updated"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
    return profile


def tune_dynamic_config(state: dict, client: Groq) -> dict:
    """AI, geçmiş döngülerin çalma oranına bakarak carry_over'ı AZALTABİLİR ama ASLA ARTIRAMAZ.

    Mantık: Bu bot tekrarı azaltmak için var. Eski playlist az dinlendiyse (kullanıcı
    sıkıldı/beğenmedi) → carry_over düşsün, daha az eski şarkı taşınsın.
    Eski playlist çok dinlendiyse → bu "daha fazla taşı" anlamına GELMEZ; çok dinlenen
    şarkılar da fazla tekrar edince sıkıcı olur. O yüzden artış hiçbir koşulda yapılmaz.
    carry_over sadece varsayılan tavanından (DEFAULT_CARRY_OVER) aşağı inebilir, geri
    yukarı çıkamaz — tek yönlü, kalıcı bir daralma eğrisi."""
    dyn = state.get("dynamic_config", _default_dynamic_config())
    archive = state.get("playlist_archive", [])[-6:]

    if len(archive) < 3:
        return dyn  # yeterli geçmiş yok, mevcut değerle devam

    history_summary = [
        {"cycle": a.get("cycle"), "score": a.get("score"), "avg_plays": a.get("avg_plays"),
         "carry_over_count": a.get("carry_over_count")}
        for a in archive
    ]

    prompt = f"""Sen bir playlist strateji uzmanısın. Bu botun amacı kullanıcının hep aynı
şarkıları dinlemekten sıkılmasını önlemek. Bu yüzden carry_over (eski playlistten yeni
playliste taşınacak şarkı sayısı) parametresi SADECE AZALTILABİLİR, asla artırılamaz.

## Mevcut Değer
carry_over: {dyn['carry_over']} (mutlak alt sınır: {CARRY_OVER_MIN}, bu döngüde çıkabileceği en yüksek değer: {dyn['carry_over']} — yani sadece eşit kalabilir ya da düşebilir)

## Son Döngüler (avg_plays = eski playlist'teki şarkıların ortalama kaç kez çalındığı)
{json.dumps(history_summary, ensure_ascii=False)}

Görev: Eğer son döngülerde avg_plays veya skor düşükse (kullanıcı eski playlist'i az
dinledi/beğenmedi) carry_over'ı 1 azaltmayı öner. Eğer avg_plays YÜKSEKSE bu bir artış
sebebi DEĞİLDİR — çok dinlenmiş olması bile aynı şarkıların tekrar tekrar gelmesini
haklı çıkarmaz, bu durumda mevcut değeri koru, asla artırma. Sadece düşüş öner ya da
"değişiklik yok" de.

SADECE JSON:
{{"carry_over": {dyn['carry_over']}, "reason": "kısa Türkçe gerekçe"}}"""

    try:
        resp = client.chat.completions.create(
            model="llama-3.3-70b-versatile", max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        parsed = _parse_ai_json(resp.choices[0].message.content)

        proposed = int(parsed.get("carry_over", dyn["carry_over"]))

        # TEK YÖNLÜ KURAL: AI ne önerirse önersin, kod seviyesinde artış kesinlikle reddedilir.
        # Sadece eşit kalabilir veya en fazla 1 birim azalabilir.
        new_carry = dyn["carry_over"] if proposed >= dyn["carry_over"] else dyn["carry_over"] - 1
        new_carry = max(CARRY_OVER_MIN, min(new_carry, PLAYLIST_SIZE))

        if new_carry != dyn["carry_over"]:
            log.info(
                f"Dynamic config güncellendi (tek yönlü, sadece azalış): "
                f"carry_over {dyn['carry_over']}→{new_carry} | Gerekçe: {parsed.get('reason', '')}"
            )
        elif proposed > dyn["carry_over"]:
            log.info(f"AI artış önerdi ({dyn['carry_over']}→{proposed}) ama tek yönlü kural gereği reddedildi, değer korunuyor.")

        dyn["carry_over"] = new_carry
        dyn["last_tuned_cycle"] = state.get("cycle", 0)
        dyn["tune_reason"] = parsed.get("reason", "")
    except Exception as e:
        log.warning(f"Dynamic config tuning başarısız, mevcut değer korunuyor: {e}")

    return dyn


def _parse_ai_json(raw: str) -> dict:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def synthesize_user_profile(state: dict, listening_data: dict, client: Groq) -> dict:
    profile = state.get("user_profile", _default_user_profile())
    archive = state.get("playlist_archive", [])[-5:]
    fingerprint = build_listening_fingerprint(listening_data)

    if not archive:
        return profile

    prompt = f"""Sen bir müzik tercihi analiz uzmanısın. Kullanıcının dinleme geçmişini inceleyip öğrenilen örüntüleri çıkar.

## Mevcut Profil
{json.dumps(profile.get("adaptation_metrics", {}), ensure_ascii=False)}

## Son 5 Döngü Arşivi
{json.dumps(archive, ensure_ascii=False)}

## Bu Dönemin Dinleme Özeti
{json.dumps(fingerprint, ensure_ascii=False)}

## Mevcut Öğrenilen Örüntüler
{json.dumps(profile.get("learned_patterns", []), ensure_ascii=False)}

Görev: Kullanıcıyı daha iyi anlamak için yeni örüntüler keşfet. Mevcut örüntüleri koru, yeni olanları ekle (toplam max 8 madde).

SADECE JSON:
{{"learned_patterns": ["örüntü1", "örüntü2"], "next_cycle_advice": "sonraki döngü için tavsiye", "confidence": 0.72}}"""

    try:
        resp = client.chat.completions.create(
            model="llama-3.3-70b-versatile", max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        parsed = _parse_ai_json(resp.choices[0].message.content)
        existing = profile.get("learned_patterns", [])
        new_patterns = parsed.get("learned_patterns", [])
        merged = list(existing)
        for p in new_patterns:
            if p not in merged:
                merged.append(p)
        profile["learned_patterns"] = merged[-8:]
        profile["next_advice"] = parsed.get("next_cycle_advice", profile.get("next_advice", ""))
        profile["adaptation_metrics"]["confidence"] = float(parsed.get("confidence", 0.0))
        log.info(f"Profil sentezi: {len(profile['learned_patterns'])} örüntü, güven %{profile['adaptation_metrics']['confidence']*100:.0f}")
    except Exception as e:
        log.warning(f"Profil sentezi başarısız: {e}")

    return profile


# ── AI ─────────────────────────────────────────────────────────────────────

def analyze_mood(listening_data, client):
    all_recent = listening_data.get("recent_tracks", [])
    prompt = f"""Sen bir müzik psikolojisi uzmanısın. Son 3 günlük dinleme verisine bakarak ruh halini analiz et.

## Son 3 Günde Dinlenen Şarkılar
{json.dumps(all_recent, ensure_ascii=False)}

## Kısa Dönem En Çok Dinlenenler
{json.dumps(listening_data.get('top_short', [])[:15], ensure_ascii=False)}

## Favori Sanatçılar
{json.dumps(listening_data.get('top_artists', [])[:8], ensure_ascii=False)}

SADECE JSON döndür:
{{"mood": "Ana ruh hali (Türkçe)", "mood_emoji": "emoji", "energy_level": "düşük/orta/yüksek",
"dominant_genres": ["tür1","tür2"], "top_artists_this_period": ["sanatçı1","sanatçı2","sanatçı3"],
"summary": "2-3 cümlelik Türkçe özet", "track_count": {len(all_recent)}}}"""

    try:
        resp = client.chat.completions.create(
            model="llama-3.3-70b-versatile", max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        return _parse_ai_json(resp.choices[0].message.content)
    except Exception as e:
        log.warning(f"Ruh hali analizi başarısız: {e}")
        return {"mood": "Bilinmiyor", "mood_emoji": "🎵", "energy_level": "orta",
                "dominant_genres": [], "top_artists_this_period": [],
                "summary": "Analiz yapılamadı.", "track_count": len(all_recent)}


def _build_learning_context(state: dict) -> str:
    profile = state.get("user_profile", {})
    archive = state.get("playlist_archive", [])[-3:]
    metrics = profile.get("adaptation_metrics", {})
    patterns = profile.get("learned_patterns", [])
    advice = profile.get("next_advice", "")

    archive_summary = [
        {"cycle": a.get("cycle"), "mood": a.get("mood"), "score": a.get("score"),
         "avg_plays": a.get("avg_plays"), "top_artists": a.get("top_artists", [])[:3]}
        for a in archive
    ]

    lines = [
        f"## Kullanıcı Profili (öğrenilmiş, döngü #{state.get('cycle', 0)})",
        f"Genre tercihi: {json.dumps(profile.get('genre_affinity', {}), ensure_ascii=False)}",
        f"Sanatçı tercihi: {json.dumps(dict(list(profile.get('artist_affinity', {}).items())[:8]), ensure_ascii=False)}",
        f"Mood-müzik haritası: {json.dumps(profile.get('mood_music_map', {}), ensure_ascii=False)}",
        "",
        "## Öğrenilen Örüntüler",
        "\n".join(f"- {p}" for p in patterns) if patterns else "- Henüz örüntü yok",
        "",
        "## Geçmiş Döngü Karşılaştırması",
        json.dumps(archive_summary, ensure_ascii=False),
        "",
        f"## Adaptasyon Notu",
        f"Güven: %{metrics.get('confidence', 0)*100:.0f} | İyileşme: {metrics.get('improvement_delta', 0):+.1f} | Çalma trendi: {metrics.get('avg_plays_trend', '0%')}",
        f"Tavsiye: {advice}" if advice else "",
    ]
    return "\n".join(lines)


def ai_analyze_and_build(listening_data, play_counts, state, candidate_ids, current_track_ids=None):
    client = Groq(api_key=GROQ_API_KEY)
    current_track_ids = current_track_ids or []

    dyn = state.get("dynamic_config", _default_dynamic_config())
    carry_over_limit = dyn["carry_over"]
    playlist_size = PLAYLIST_SIZE

    if current_track_ids and play_counts:
        sorted_old = sorted(play_counts.items(), key=lambda x: x[1], reverse=True)
        carry_over_pool = [tid for tid, count in sorted_old if count > 0][:carry_over_limit]
        banned_ids = [tid for tid in current_track_ids if tid not in carry_over_pool]
    else:
        carry_over_pool = []
        banned_ids = current_track_ids

    fresh_candidates = [tid for tid in candidate_ids if tid not in banned_ids]
    log.info(f"Aday: {len(fresh_candidates)} taze + {len(carry_over_pool)} taşınabilir eski şarkı")

    learning_context = _build_learning_context(state)
    mood_data = analyze_mood(listening_data, client)
    log.info(f"Ruh hali: {mood_data.get('mood')} {mood_data.get('mood_emoji')}")

    prompt = f"""Sen bir müzik küratörüsün. {playlist_size} şarkılık playlist seç.

{learning_context}

Döngü #{state['cycle'] + 1}
Ruh hali: {json.dumps(mood_data, ensure_ascii=False)}
Son dinlenenler: {json.dumps(listening_data['recent_tracks'][:30], ensure_ascii=False)}
Top şarkılar: {json.dumps(listening_data['top_short'][:20], ensure_ascii=False)}
Çalma sayıları: {json.dumps(play_counts, ensure_ascii=False)}
Top sanatçılar: {json.dumps(listening_data['top_artists'][:10], ensure_ascii=False)}

## TAZE ADAY ŞARKILAR (öncelikli olarak bunlardan seç)
{json.dumps(fresh_candidates[:80], ensure_ascii=False)}

## TAŞINABİLİR ESKİ ŞARKILAR (geçen dönem çok çalındı, en fazla {carry_over_limit} tane kullanabilirsin)
{json.dumps(carry_over_pool, ensure_ascii=False)}

## KESİNLİKLE YASAK (bu şarkıları ekleme)
{json.dumps(banned_ids[:50], ensure_ascii=False)}

Görev:
1. Öğrenilen profil ve örüntülere göre seçim yap — kullanıcıyı tanıdığını göster.
2. Taze aday şarkılardan en az {playlist_size - carry_over_limit} şarkı seç.
3. Taşınabilir eski şarkılardan en fazla {carry_over_limit} tane ekleyebilirsin.
4. Yasak listesindeki şarkıları KESİNLİKLE ekleme.
5. Ruh haline uygun seç, çeşitlilik ekle, sıralamayı karıştır.

SADECE JSON:
{{"track_ids":["id1","id2"],"score":7.5,"analysis":"Türkçe analiz","notes":"sonraki döngü notu"}}"""

    resp = client.chat.completions.create(
        model="llama-3.3-70b-versatile", max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )
    parsed = _parse_ai_json(resp.choices[0].message.content)
    track_ids = parsed.get("track_ids", [])[:playlist_size]

    carry_over_used = [tid for tid in track_ids if tid in carry_over_pool]
    if len(carry_over_used) > carry_over_limit:
        excess = set(carry_over_used[carry_over_limit:])
        track_ids = [tid for tid in track_ids if tid not in excess]
        log.info(f"Carry-over limiti: {len(excess)} fazla şarkı çıkarıldı")
    track_ids = [tid for tid in track_ids if tid not in banned_ids]

    fresh_count = len([tid for tid in track_ids if tid not in current_track_ids])
    carry_count = len([tid for tid in track_ids if tid in carry_over_pool])
    log.info(f"AI tamamlandı. Skor: {parsed.get('score')}, Taze: {fresh_count}, Eski: {carry_count}, Toplam: {len(track_ids)}")
    return track_ids, parsed.get("notes", ""), float(parsed.get("score", 5.0)), parsed.get("analysis", ""), mood_data, carry_count


# ── Ana Döngü ─────────────────────────────────────────────────────────────

def run_cycle(manual=False):
    global is_running
    if is_running:
        log.warning("Zaten çalışıyor.")
        return {"status": "already_running"}
    is_running = True
    trigger = "Manuel" if manual else "Otomatik"
    log.info(f"═══════════ Döngü Başlıyor ({trigger}) ═══════════")

    try:
        state = load_state()
        sp = get_spotify()

        playlist_id = find_or_create_playlist(sp, state)
        state["playlist_id"] = playlist_id
        save_state(state)

        current_tracks = []
        if state.get("last_update"):
            try:
                items = sp.playlist_items(playlist_id, fields="items(track(id,name,artists,album))")
                for item in items["items"]:
                    t = item["track"]
                    if t and t.get("id"):
                        current_tracks.append({
                            "id": t["id"], "name": t["name"],
                            "artist": t["artists"][0]["name"], "album": t["album"]["name"],
                        })
            except Exception as e:
                log.warning(f"Playlist okunamadı: {e}")

        listening_data = get_listening_data(sp)
        fingerprint = build_listening_fingerprint(listening_data)

        play_counts = {}
        avg_plays = 0.0
        current_ids_for_ai = [t["id"] for t in current_tracks]
        if current_tracks:
            play_counts = get_playlist_play_counts(current_ids_for_ai, listening_data["recent_tracks"])
            avg_plays = sum(play_counts.values()) / max(len(play_counts), 1)

        analyze_patterns(state, listening_data, play_counts, current_ids_for_ai)

        groq_client = Groq(api_key=GROQ_API_KEY)
        state["user_profile"] = synthesize_user_profile(state, listening_data, groq_client)
        state["dynamic_config"] = tune_dynamic_config(state, groq_client)

        candidate_ids = list({t["id"] for t in (
            listening_data["top_short"] + listening_data["top_medium"] +
            listening_data["saved_tracks"] + listening_data["recent_tracks"]
        ) if t.get("id")})

        new_track_ids, ai_notes, score, analysis, mood_data, carry_count = ai_analyze_and_build(
            listening_data, play_counts, state, candidate_ids, current_ids_for_ai,
        )

        archiving_cycle = state["cycle"]
        old_score = state["feedback_history"][-1]["score"] if state.get("feedback_history") else None
        if current_tracks:
            save_to_excel(current_tracks, archiving_cycle, old_score, play_counts)

        update_playlist(sp, playlist_id, new_track_ids)

        now = datetime.datetime.now(datetime.timezone.utc).isoformat()
        state["cycle"] += 1
        state["last_update"] = now
        state["ai_notes"] = ai_notes

        top_played = sorted(
            [{"id": tid, "name": next((t["name"] for t in current_tracks if t["id"] == tid), tid),
              "plays": c} for tid, c in play_counts.items() if c > 0],
            key=lambda x: x["plays"], reverse=True,
        )[:5]

        archive_entry = {
            "cycle": state["cycle"],
            "archived_at": now,
            "score": score,
            "avg_plays": round(avg_plays, 2),
            "mood": mood_data.get("mood", ""),
            "energy": mood_data.get("energy_level", ""),
            "track_count": len(new_track_ids),
            "top_played": top_played,
            "top_artists": fingerprint["top_artists"][:5],
            "genres": mood_data.get("dominant_genres", fingerprint["top_genres"][:5]),
            "carry_over_count": carry_count,
            "fresh_count": len(new_track_ids) - carry_count,
            "listening_fingerprint": fingerprint,
        }
        if "playlist_archive" not in state:
            state["playlist_archive"] = []
        state["playlist_archive"].append(archive_entry)
        state["playlist_archive"] = state["playlist_archive"][-ARCHIVE_LIMIT:]

        save_cycle_summary_to_excel(
            state["cycle"], mood_data, score, len(new_track_ids),
            avg_plays, carry_count, analysis,
        )

        if "mood_history" not in state:
            state["mood_history"] = []
        state["mood_history"].append({"date": now, "cycle": state["cycle"], "trigger": trigger, **mood_data})

        if "feedback_history" not in state:
            state["feedback_history"] = []
        state["feedback_history"].append({
            "cycle": state["cycle"], "date": now,
            "score": score, "avg_plays": avg_plays, "analysis": analysis, "notes": ai_notes,
        })

        save_state(state)
        log.info(f"═══════════ Döngü #{state['cycle']} Tamamlandı ═══════════")
        return {"status": "ok", "cycle": state["cycle"], "tracks": len(new_track_ids), "mood": mood_data}

    except Exception as e:
        log.error(f"Döngü hatası: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}
    finally:
        is_running = False


# ── Flask API ──────────────────────────────────────────────────────────────

def check_auth(req):
    secret = req.headers.get("X-API-Secret") or req.args.get("secret")
    return secret == API_SECRET


@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "ok", "bot": "Spotify AI Playlist Bot"})


@app.route("/status", methods=["GET"])
def status():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    state = load_state()
    return jsonify({
        "cycle": state.get("cycle", 0),
        "last_update": state.get("last_update"),
        "playlist_id": state.get("playlist_id"),
        "is_running": is_running,
        "feedback_history": state.get("feedback_history", []),
        "mood_history": state.get("mood_history", []),
        "user_profile": state.get("user_profile", {}),
        "dynamic_config": state.get("dynamic_config", {}),
        "playlist_archive_summary": state.get("playlist_archive", [])[-10:],
    })


@app.route("/history", methods=["GET"])
def history():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    state = load_state()
    return jsonify(state.get("playlist_archive", []))


@app.route("/history/excel", methods=["GET"])
def history_excel():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    if not os.path.exists(HISTORY_FILE):
        return jsonify({"error": "Henüz Excel geçmişi yok"}), 404
    return send_file(HISTORY_FILE, as_attachment=True, download_name="playlist_history.xlsx")


@app.route("/trigger", methods=["POST"])
def trigger():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    if is_running:
        return jsonify({"status": "already_running"}), 409
    t = threading.Thread(target=lambda: run_cycle(manual=True), daemon=True)
    t.start()
    return jsonify({"status": "started"})


@app.route("/mood-history", methods=["GET"])
def mood_history():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    state = load_state()
    return jsonify(state.get("mood_history", []))


# ── Başlat ─────────────────────────────────────────────────────────────────

def should_run_on_startup(state: dict) -> bool:
    if not state.get("last_update"):
        return True
    try:
        last = datetime.datetime.fromisoformat(state["last_update"])
        if last.tzinfo is None:
            last = last.replace(tzinfo=datetime.timezone.utc)
        now = datetime.datetime.now(datetime.timezone.utc)
        return (now - last) > datetime.timedelta(hours=47)
    except Exception:
        return True


def scheduler_loop():
    schedule.every(48).hours.do(run_cycle)
    while True:
        schedule.run_pending()
        time.sleep(60)


def main():
    log.info("Spotify AI Bot başlatıldı.")
    state = load_state()
    if should_run_on_startup(state):
        log.info("Startup döngüsü başlatılıyor.")
        threading.Thread(target=run_cycle, daemon=True).start()
    else:
        log.info("Son güncelleme 47 saatten yakın, startup döngüsü atlandı.")
    threading.Thread(target=scheduler_loop, daemon=True).start()
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)


if __name__ == "__main__":
    main()
